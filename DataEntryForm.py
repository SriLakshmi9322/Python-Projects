from tkinter import Tk, PhotoImage, Label, StringVar, Entry, Text, Button, END
from tkinter import messagebox
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib

root = Tk()
root.title("Data Entry")
root.geometry("700x400+300+200")
root.resizable(False, False)
root.configure(bg="#326273")

file = pathlib.Path('Backend_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"

    file.save('Backend_data.xlsx')


def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combbox.get()
    address = addressEntry.get(1.0, END)

    file = openpyxl.load_workbook('Backend_data.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)

    file.save(r'Backend_data.xlsx')

    messagebox.showinfo('info', 'Detail Added!')

    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0, END)


def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0, END)


# Icon
icon_image = PhotoImage(file="Images/data_entry_logo.png")
root.iconphoto(False, icon_image)

# Heading
Label(root, text="Please fill out this Entry Form: ", font="arial 13",
      bg="#326273", fg="#fff").place(x=20, y=20)

# Label
Label(root, text="Name", font=23, bg="#326273", fg="#fff").place(x=50,
                                                                 y=100)
Label(root, text="Contact No.", font=23, bg="#326273", fg="#fff").place(x=50,
                                                                        y=150)
Label(root, text="Age", font=23, bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text="Gender", font=23, bg="#326273", fg="#fff").place(x=380,
                                                                   y=200)
Label(root, text="Address", font=23, bg="#326273", fg="#fff").place(x=50,
                                                                    y=250)

# Entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=40, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=40, bd=2, font=20)
ageEntry = Entry(root, textvariable=AgeValue, width=14, bd=2, font=20)

# Gender
gender_combbox = Combobox(root, values=['Male', 'Female'], font="arial 14",
                          state='r', width=14)
gender_combbox.place(x=470, y=200)
gender_combbox.set('Male')

addressEntry = Text(root, width=55, height=4, bd=4)

nameEntry.place(x=200, y=100)
contactEntry.place(x=200, y=150)
ageEntry.place(x=200, y=200)
addressEntry.place(x=200, y=250)

Button(root, text="Submit", bg="#326273", fg="white", width=15,
       height=2, command=submit).place(x=200, y=340)
Button(root, text="Clear", bg="#326273", fg="white", width=15,
       height=2, command=clear).place(x=365, y=340)
Button(root, text="Exit", bg="#326273", fg="white", width=15,
       height=2, command=lambda: root.destroy()).place(x=530, y=340)

root.mainloop()
